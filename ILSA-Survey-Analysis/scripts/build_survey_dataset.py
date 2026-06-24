"""
Build ILSA Survey CLEAN Excel dataset from JSON files.
Mirrors the structure of ILSA_Meta_Analysis_Dataset_CLEAN.xlsx.
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ── Paths ──────────────────────────────────────────────────────────────────────
JSON_DIR = Path("/Users/mrved/Desktop/ILSA Survey/ilsa_survey_articles_json/json")
OUTPUT_DIR = Path("/Users/mrved/Desktop/ILSA Survey/ILSA-Survey-Analysis/outputs")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
OUTPUT_PATH = OUTPUT_DIR / "ILSA_Survey_Dataset_CLEAN.xlsx"

# Articles excluded from survey scope
EXCLUDED_FILE_NAMES = {
    "34. Miranda et al. (2021). Development of INSVAGRAM: An English Subject-Verb Agreement Mobile Learning Application.pdf",
    "58. Rodriguez-Barrios et al. (2021). Bayesian Approach to Analyze Reading Comprehension: A Case Study in Elementary School Children in Mexico.pdf",
    "20. Bezek-Güre et al. (2020) Analysis of Factors Effecting PISA 2015 Mathematics Literacy via Educational Data Mining TR.pdf",
    # Duplicate: same paper as 81. Schipper et al.
    "123. Schipper et al. (2025). Identifying students’ solution strategies in digital mathematics assessment using log data.pdf",
}

# ── Taxonomy (copied from ILSA_LLMs academic_taxonomy.py) ────────────────────

STUDY_EMPIRICAL_ML = "Empirical Study - Machine Learning"
STUDY_EMPIRICAL_TRADITIONAL = "Empirical Study - Traditional Statistics"
STUDY_TECHNICAL_FRAMEWORK = "Technical/Assessment Framework"
STUDY_DESCRIPTIVE_NATIONAL = "Descriptive National Report"
STUDY_UNCLASSIFIED = "Unclassified"

_SENTINEL_PREFIXES = ("n/a:", "not reported", "missing")
_NATIONAL_REPORT_MARKERS = (
    "national report", "international report", "initial findings",
    "highlights", "volume i", "volume ii", "volume iii",
    "compass brief", "policy brief",
)
_FALSE_POSITIVE_PHRASES = ("self-reported", "self reported", "technical efficiency", "technical (stem) track", "technical engineering problem")

_REPORT_SIGNAL_RE = re.compile(
    r"\b(?:technical\s+report|national\s+report|international\s+report|"
    r"assessment\s+framework|user\s+guide|methods\s+and\s+procedures|"
    r"encyclopedia|idb\s+user)\b",
    re.IGNORECASE,
)
_TECHNICAL_WORD_RE = re.compile(r"\btechnical\b", re.IGNORECASE)
_REPORT_WORD_RE = re.compile(r"\breport\b", re.IGNORECASE)

_SCHEMA_CATEGORY_MAP: dict[str, tuple[str, str]] = {
    "socioeconomic": ("Student Level", "Student: SES"),
    "demographic": ("Student Level", "Student: Demographic"),
    "student_attitude": ("Student Level", "Student: Attitudinal/Behavioral"),
    "student_behavior": ("Student Level", "Student: Attitudinal/Behavioral"),
    "prior_achievement": ("Student Level", "Student: Prior Achievement"),
    "process_data": ("Student Level", "Student: Process Data"),
    "teacher": ("School/Teacher Level", "School/Teacher: Practice"),
    "school": ("School/Teacher Level", "School/Teacher: Context"),
    "ict": ("Student Level", "Student: Attitudinal/Behavioral"),
    "curriculum": ("School/Teacher Level", "School/Teacher: Context"),
    "parent_home": ("Student Level", "Student: SES"),
    "peer_effects": ("School/Teacher Level", "School/Teacher: Context"),
    "system_level": ("System/Country Level", "System: Policy/Context"),
}

_TARGET_RULES: list[tuple[re.Pattern, str, str]] = [
    (re.compile(r"pv\d*\s*math|pv\d*math|mathem|numeracy|quantitative literacy", re.I), "Mathematics", "Cognitive Achievement"),
    (re.compile(r"pv\d*\s*scie|pv\d*scie|science achievement|scientific literacy", re.I), "Science", "Cognitive Achievement"),
    (re.compile(r"pv\d*\s*read|pv\d*read|reading literacy|reading achievement", re.I), "Reading", "Cognitive Achievement"),
    (re.compile(r"computer.?based|digital literacy|ict literacy|computational", re.I), "Digital/Computer Literacy", "Cognitive Achievement"),
    (re.compile(r"civic|citizenship|global competenc", re.I), "Civic Education", "Cognitive Achievement"),
    (re.compile(r"problem.?solv|ps task|creative thinking", re.I), "Problem Solving", "Cognitive Achievement"),
    (re.compile(r"response time|log.?file|process data|votat|time on task|click", re.I), "Non-Cognitive / Process Output", "Process Data / Log Metrics"),
    (re.compile(r"dropout|well.?being|anxiety|motivation|self.?efficacy|belonging|climate", re.I), "Non-Cognitive / Process Output", "Attitudinal / Affective"),
    (re.compile(r"resilien|policy|leadership|instructional time", re.I), "Non-Cognitive / Process Output", "Policy / System Outcome"),
    (re.compile(r"achievement|proficiency|literacy|score|performance", re.I), "Composite / Multi-Domain", "Cognitive Achievement"),
]

_ML_FAMILY_RULES: list[tuple[re.Pattern, str]] = [
    (re.compile(r"random forest|xgboost|gradient boost|lightgbm|catboost|decision tree|adaboost|bagging|ensemble|extra trees", re.I), "Tree-Based / Ensemble Learning"),
    (re.compile(r"logistic regression|linear regression|lasso|ridge|elastic net|glm|probit", re.I), "Generalized Linear Models (GLM)"),
    (re.compile(r"neural|deep learning|cnn|lstm|rnn|mlp|transformer", re.I), "Deep Learning"),
    (re.compile(r"k-means|kmeans|dbscan|cluster|latent class|latent profile|topic model", re.I), "Clustering / Unsupervised Learning"),
    (re.compile(r"\birt\b|mplus|hlm|sem|structural equation|multilevel|mixed effect|conquest|winsteps|tam\b|lavaan|plausible value|rubin", re.I), "Traditional Psychometrics / Multilevel Modeling"),
    (re.compile(r"svm|support vector|naive bayes|knn|k-nearest", re.I), "Other ML / Not Classified"),
]

_PV_LABEL_MAP: dict[str, str] = {
    "rubin_rules": "Pooled PVs (Rubin Rules)",
    "single_pv": "Single PV Draw",
    "average_pv": "Average PVs",
    "all_pv": "All PVs Analyzed Separately",
    "mitml": "Multiple Imputation",
    "wle": "WLE / IRT Theta",
    "irt_theta": "WLE / IRT Theta",
    "not_applicable": "Not Applicable (Framework)",
    "not_reported": "Not Reported",
}

_MD_LABEL_MAP: dict[str, str] = {
    "listwise_deletion": "Listwise Deletion",
    "pairwise_deletion": "Pairwise Deletion",
    "mean_imputation": "Mean Imputation",
    "single_imputation": "Single Imputation",
    "knn_imputation": "KNN Imputation",
    "multiple_imputation": "Multiple Imputation",
    "not_reported": "Not Reported",
}


def _norm(v: Any) -> str:
    if v is None or (isinstance(v, float) and str(v) == "nan"):
        return ""
    return str(v).strip()


def _is_sentinel(v: Any) -> bool:
    t = _norm(v).lower()
    if not t:
        return True
    return any(t.startswith(p) for p in _SENTINEL_PREFIXES)


def _norm_pub(v: Any) -> str:
    return _norm(v).lower().replace("-", "_").replace(" ", "_")


def _mask_fp(text: str) -> str:
    masked = text.lower()
    for p in _FALSE_POSITIVE_PHRASES:
        masked = masked.replace(p, " ")
    return masked


def _is_empirical_channel(row: dict) -> bool:
    pub = _norm_pub(row.get("publication_type"))
    return pub in ("journal", "conference")


def _has_ml(row: dict) -> bool:
    for col in ("ml_techniques", "ml_primary", "ml_all_techniques"):
        text = _norm(row.get(col)).lower()
        if not text or _is_sentinel(text):
            continue
        if text == "present":
            continue
        if re.search(r"\b(irt|mplus|hlm|sem|structural equation|multilevel model)\b", text, re.I) and \
           not re.search(r"random forest|xgboost|neural|svm|deep learning|classif", text, re.I):
            continue
        return True
    return False


def assign_document_class(row: dict) -> str:
    if _is_empirical_channel(row):
        return "empirical_article"
    sc = _norm(row.get("source_category")).lower()
    pub = _norm_pub(row.get("publication_type"))
    title_masked = _mask_fp(_norm(row.get("title")))
    combined = f"{title_masked} {_mask_fp(sc)} {pub}"
    if _REPORT_SIGNAL_RE.search(combined):
        return "technical_report"
    if _TECHNICAL_WORD_RE.search(title_masked) and pub in ("report", "book_chapter", ""):
        return "technical_report"
    if _REPORT_WORD_RE.search(title_masked) and "self" not in title_masked:
        return "technical_report"
    if pub == "report" or sc == "technical_report":
        return "technical_report"
    if sc == "methodology_paper" and pub in ("report", "book_chapter"):
        return "technical_report"
    if "peer_reviewed" in pub or pub == "journal":
        return "empirical_article"
    return "unclassified"


def classify_study_filter(row: dict) -> str:
    pub = _norm_pub(row.get("publication_type"))
    sc = _norm(row.get("source_category")).lower()

    if _is_empirical_channel(row):
        return STUDY_EMPIRICAL_ML if _has_ml(row) else STUDY_EMPIRICAL_TRADITIONAL

    if sc == "methodology_paper":
        return STUDY_TECHNICAL_FRAMEWORK

    doc_class = assign_document_class(row)
    if doc_class == "empirical_article":
        return STUDY_EMPIRICAL_ML if _has_ml(row) else STUDY_EMPIRICAL_TRADITIONAL

    if doc_class == "technical_report" or pub == "report":
        blob = " ".join(_norm(row.get(k)) for k in ("title", "source_category", "outcome_summary")).lower()
        if any(m in blob for m in _NATIONAL_REPORT_MARKERS):
            return STUDY_DESCRIPTIVE_NATIONAL
        return STUDY_TECHNICAL_FRAMEWORK

    if pub in ("thesis", "preprint", "book_chapter"):
        return STUDY_EMPIRICAL_ML if _has_ml(row) else STUDY_EMPIRICAL_TRADITIONAL

    return STUDY_UNCLASSIFIED


def map_ml_family(ml_text: Any, study_filter: str = "") -> str:
    if study_filter in (STUDY_TECHNICAL_FRAMEWORK, STUDY_DESCRIPTIVE_NATIONAL):
        return "N/A: Technical Report"
    text = _norm(ml_text)
    if _is_sentinel(text):
        return "Not Reported: Likely Traditional Methods"
    for pattern, family in _ML_FAMILY_RULES:
        if pattern.search(text):
            return family
    return "Other ML / Not Classified"


def map_pv_filter(value: Any) -> str:
    key = _norm(value).lower().replace(" ", "_")
    if _is_sentinel(value):
        return "Not Reported"
    return _PV_LABEL_MAP.get(key, "Other")


def map_md_filter(value: Any) -> str:
    key = _norm(value).lower().replace(" ", "_")
    if _is_sentinel(value):
        return "Not Reported"
    return _MD_LABEL_MAP.get(key, "Other")


def map_weights_filter(value: Any) -> str:
    if value is True or _norm(value).upper() == "TRUE":
        return "True"
    if value is False or _norm(value).upper() == "FALSE":
        return "False"
    return "Unknown"


def map_target_domain_dimension(target_variable: Any, study_filter: str = "") -> tuple[str, str]:
    if study_filter in (STUDY_TECHNICAL_FRAMEWORK, STUDY_DESCRIPTIVE_NATIONAL):
        return "N/A: Technical Report", "Methodological (no DV)"
    if _is_sentinel(target_variable):
        return "Other / Unspecified", "Other"
    text = _norm(target_variable)
    for pattern, domain, dimension in _TARGET_RULES:
        if pattern.search(text):
            return domain, dimension
    return "Other / Unspecified", "Other"


def categorize_predictor(variable_name: Any, schema_category: Any = None, study_filter: str = "") -> tuple[str, str]:
    if study_filter in (STUDY_TECHNICAL_FRAMEWORK, STUDY_DESCRIPTIVE_NATIONAL):
        return "N/A: Technical Report", "N/A: Technical Report"
    sc = _norm(schema_category).lower()
    if sc in _SCHEMA_CATEGORY_MAP:
        return _SCHEMA_CATEGORY_MAP[sc]
    text = _norm(variable_name).lower()
    if _is_sentinel(text):
        return "Unspecified", "Other"
    if any(w in text for w in ("escs", "hisei", "pared", "ses", "homepos", "wealth", "income", "books")):
        return "Student Level", "Student: SES"
    if any(w in text for w in ("gender", "sex", "immig", "age", "grade", "birth")):
        return "Student Level", "Student: Demographic"
    if any(w in text for w in ("prior", "achievement", "wle", "pv1", "pv2", "score")):
        return "Student Level", "Student: Prior Achievement"
    if any(w in text for w in ("time on task", "log", "process", "votat", "ict use", "click")):
        return "Student Level", "Student: Process Data"
    if any(w in text for w in ("teacher", "instruction", "pedagog")):
        return "School/Teacher Level", "School/Teacher: Practice"
    if any(w in text for w in ("school", "class size", "climate", "urban", "rural", "private", "public")):
        return "School/Teacher Level", "School/Teacher: Context"
    if any(w in text for w in ("gdp", "gini", "country", "policy", "system", "tracking")):
        return "System/Country Level", "System: Policy/Context"
    if any(w in text for w in ("motiv", "anxiety", "efficacy", "interest", "belong", "enjoy")):
        return "Student Level", "Student: Attitudinal/Behavioral"
    return "Student Level", "Student: Attitudinal/Behavioral"


# ── JSON extraction ───────────────────────────────────────────────────────────

def load_json_files(json_dir: Path) -> list[dict]:
    records = []
    for path in sorted(json_dir.glob("*.json")):
        try:
            with open(path, encoding="utf-8") as f:
                records.append(json.load(f))
        except Exception as e:
            print(f"  WARN: {path.name}: {e}", file=sys.stderr)
    return records


def extract_master_row(rec: dict, json_path: str) -> dict:
    meta = rec.get("metadata", {})
    data = rec.get("data", {})
    sd = data.get("survey_design", {})
    sample = data.get("sample_details", {})
    ml = data.get("ml_techniques", {})

    authors = meta.get("authors", [])
    authors_str = "; ".join(authors) if isinstance(authors, list) else _norm(authors)

    countries = sample.get("countries", [])
    if isinstance(countries, list):
        codes = [c.get("country_code", "") for c in countries if isinstance(c, dict)]
        countries_formatted = "; ".join(c for c in codes if c)
        countries_json = json.dumps(countries)
    else:
        countries_formatted = ""
        countries_json = "[]"

    ml_primary = ml.get("primary") if isinstance(ml, dict) else None
    ml_all = ml.get("all_techniques", []) if isinstance(ml, dict) else []
    ml_all_str = "; ".join(ml_all) if isinstance(ml_all, list) else _norm(ml_all)

    confounders = data.get("confounders_identified", [])
    confounder_names = "; ".join(
        c.get("variable_name", "") for c in confounders if isinstance(c, dict)
    ) if isinstance(confounders, list) else ""

    findings = data.get("main_findings", [])
    primary_finding = findings[0].get("standardized_conclusion", "") if findings and isinstance(findings[0], dict) else ""
    effect_sizes = [f.get("performance_metrics", "") for f in findings if isinstance(f, dict) and f.get("performance_metrics")]
    effect_size = effect_sizes[0] if effect_sizes else ""

    row: dict = {
        "file_name": meta.get("file_name", ""),
        "doi": meta.get("doi", ""),
        "title": meta.get("title", ""),
        "authors": authors_str,
        "year": meta.get("year"),
        "publication_type": meta.get("publication_type", ""),
        "source_category": meta.get("source_category", ""),
        "venue": meta.get("venue", ""),
        "open_access": meta.get("open_access"),
        "corpus_source": "ilsa_survey",
        "json_source_path": json_path,
        "student_weights_used": sd.get("student_weights_used"),
        "replicate_weights_used": sd.get("replicate_weights_used"),
        "weight_variable_name": sd.get("weight_variable_name"),
        "weight_fields_interpretation": sd.get("weight_fields_interpretation"),
        "plausible_values_handling": data.get("plausible_values_handling"),
        "missing_data_handling": data.get("missing_data_handling"),
        "handling_not_reported_explanation": data.get("handling_not_reported_explanation"),
        "research_design_type": data.get("research_design_type"),
        "outcome_summary": data.get("outcome_summary"),
        "null_fields_interpretation": data.get("null_fields_interpretation"),
        "ml_primary": ml_primary,
        "ml_all_techniques": ml_all_str,
        "ml_techniques": ml_all_str or _norm(ml_primary),
        "total_students": sample.get("total_students"),
        "sample_filtering_criteria": sample.get("sample_filtering_criteria"),
        "countries_formatted": countries_formatted,
        "countries_json": countries_json,
        "effect_size": effect_size,
        "primary_finding": primary_finding,
        "sample_size": sample.get("total_students"),
        "confounders": confounder_names,
    }
    return row


def extract_findings_rows(rec: dict) -> list[dict]:
    meta = rec.get("metadata", {})
    file_name = meta.get("file_name", "")
    doi = meta.get("doi", "")
    pub_type = meta.get("publication_type", "")
    source_cat = meta.get("source_category", "")

    rows = []
    for finding in rec.get("data", {}).get("main_findings", []):
        if not isinstance(finding, dict):
            continue
        top_predictors = finding.get("top_predictors", [])
        if isinstance(top_predictors, list):
            top_predictors_str = "; ".join(str(p) for p in top_predictors)
        else:
            top_predictors_str = _norm(top_predictors)

        rows.append({
            "file_name": file_name,
            "doi": doi,
            "dataset_used": finding.get("dataset_used", ""),
            "target_variable": finding.get("target_variable", ""),
            "top_predictors": top_predictors_str,
            "performance_metrics": finding.get("performance_metrics", ""),
            "standardized_conclusion": finding.get("standardized_conclusion", ""),
            "publication_type": pub_type,
            "source_category": source_cat,
            "effect_size": finding.get("performance_metrics", ""),
            "primary_finding": finding.get("standardized_conclusion", ""),
        })
    return rows


def extract_confounder_rows(rec: dict) -> list[dict]:
    meta = rec.get("metadata", {})
    file_name = meta.get("file_name", "")
    doi = meta.get("doi", "")
    pub_type = meta.get("publication_type", "")
    source_cat = meta.get("source_category", "")

    rows = []
    for conf in rec.get("data", {}).get("confounders_identified", []):
        if not isinstance(conf, dict):
            continue
        rows.append({
            "file_name": file_name,
            "doi": doi,
            "variable_code": conf.get("variable_code", ""),
            "variable_name": conf.get("variable_name", ""),
            "category": conf.get("category", ""),
            "publication_type": pub_type,
            "source_category": source_cat,
        })
    return rows


# ── Dashboard ─────────────────────────────────────────────────────────────────

def build_dashboard(df_master: pd.DataFrame, df_findings: pd.DataFrame, df_confounders: pd.DataFrame) -> pd.DataFrame:
    rows = [
        {"Section": "Overview", "Dimension": "Metric", "Value": "Count", "Notes": ""},
        {"Section": "Overview", "Dimension": "Unique articles (Master)", "Value": len(df_master), "Notes": ""},
        {"Section": "Overview", "Dimension": "Main findings rows", "Value": len(df_findings), "Notes": ""},
        {"Section": "Overview", "Dimension": "Confounder rows", "Value": len(df_confounders), "Notes": ""},
    ]

    def _add(section: str, series: pd.Series) -> None:
        for label, count in series.value_counts().items():
            rows.append({"Section": section, "Dimension": str(label), "Value": int(count), "Notes": "Use as Excel column filter"})

    if "study_filter_type" in df_master.columns:
        _add("Study filter (Master)", df_master["study_filter_type"])
    if "ml_family" in df_master.columns:
        _add("ML family (Master)", df_master["ml_family"])
    if "target_domain" in df_findings.columns:
        _add("Target domain (Findings)", df_findings["target_domain"])
    if "predictor_category" in df_confounders.columns:
        _add("Predictor category (Confounders)", df_confounders["predictor_category"])

    return pd.DataFrame(rows)


# ── Excel formatting ──────────────────────────────────────────────────────────

def _add_table(ws, table_name: str) -> None:
    if ws.max_row < 2:
        return
    ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    for name in list(ws.tables.keys()):
        del ws.tables[name]
    t = Table(displayName=table_name, ref=ref)
    t.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(t)


def _freeze_and_width(ws) -> None:
    ws.freeze_panes = "A2"
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        header = ws.cell(1, col_idx).value
        max_len = len(str(header or ""))
        for row_idx in range(2, min(ws.max_row + 1, 500)):
            val = ws.cell(row_idx, col_idx).value
            if val is not None:
                max_len = max(max_len, min(len(str(val)), 52))
        ws.column_dimensions[letter].width = max(10, min(max_len + 2, 52))


def format_sheet(ws, table_name: str) -> None:
    _freeze_and_width(ws)
    _add_table(ws, table_name)


def write_df_to_sheet(ws, df: pd.DataFrame) -> None:
    ws.append(list(df.columns))
    for row in df.itertuples(index=False):
        ws.append(list(row))


# ── Main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    print(f"Loading JSON files from {JSON_DIR}")
    records = load_json_files(JSON_DIR)
    print(f"  Loaded {len(records)} files")

    master_rows = []
    findings_rows = []
    confounder_rows = []

    for rec in records:
        meta = rec.get("metadata", {})
        json_path = meta.get("file_name", "")
        if json_path in EXCLUDED_FILE_NAMES:
            continue
        master_rows.append(extract_master_row(rec, json_path))
        findings_rows.extend(extract_findings_rows(rec))
        confounder_rows.extend(extract_confounder_rows(rec))

    df_master = pd.DataFrame(master_rows)
    df_findings = pd.DataFrame(findings_rows)
    df_confounders = pd.DataFrame(confounder_rows)

    print(f"  Master: {len(df_master)} rows, Findings: {len(df_findings)} rows, Confounders: {len(df_confounders)} rows")

    # Apply taxonomy
    print("Applying taxonomy...")

    df_master["document_class"] = df_master.apply(lambda r: assign_document_class(r.to_dict()), axis=1)
    df_master["study_filter_type"] = df_master.apply(lambda r: classify_study_filter(r.to_dict()), axis=1)
    df_master["ml_family"] = [
        map_ml_family(df_master["ml_techniques"].iat[i], df_master["study_filter_type"].iat[i])
        for i in range(len(df_master))
    ]
    df_master["pv_filter_label"] = df_master["plausible_values_handling"].map(map_pv_filter)
    df_master["md_filter_label"] = df_master["missing_data_handling"].map(map_md_filter)
    df_master["weights_filter"] = df_master["student_weights_used"].map(map_weights_filter)

    # Merge study_filter into findings & confounders
    lookup = df_master[["file_name", "study_filter_type", "document_class"]].drop_duplicates("file_name")

    if not df_findings.empty:
        df_findings = df_findings.merge(lookup, on="file_name", how="left")
        domains, dimensions, pred_cats = [], [], []
        for i in range(len(df_findings)):
            sf = _norm(df_findings["study_filter_type"].iat[i]) if "study_filter_type" in df_findings.columns else ""
            dom, dim = map_target_domain_dimension(df_findings["target_variable"].iat[i], sf)
            domains.append(dom)
            dimensions.append(dim)
            # predictor filter categories from top_predictors
            preds_text = _norm(df_findings["top_predictors"].iat[i]) if "top_predictors" in df_findings.columns else ""
            cats = []
            if preds_text and not _is_sentinel(preds_text):
                for p in re.split(r"[;,|]", preds_text):
                    p = p.strip()
                    if p:
                        _, cat = categorize_predictor(p, study_filter=sf)
                        if cat not in cats:
                            cats.append(cat)
            pred_cats.append("; ".join(cats))
        df_findings["target_domain"] = domains
        df_findings["target_dimension"] = dimensions
        df_findings["predictor_filter_categories"] = pred_cats

    if not df_confounders.empty:
        df_confounders = df_confounders.merge(lookup[["file_name", "study_filter_type"]], on="file_name", how="left")
        levels, categories = [], []
        for i in range(len(df_confounders)):
            sf = _norm(df_confounders["study_filter_type"].iat[i]) if "study_filter_type" in df_confounders.columns else ""
            vn = df_confounders["variable_name"].iat[i] if "variable_name" in df_confounders.columns else ""
            sc = df_confounders["category"].iat[i] if "category" in df_confounders.columns else None
            lvl, cat = categorize_predictor(vn, sc, study_filter=sf)
            levels.append(lvl)
            categories.append(cat)
        df_confounders["predictor_level"] = levels
        df_confounders["predictor_category"] = categories

    df_dashboard = build_dashboard(df_master, df_findings, df_confounders)

    # Reorder master columns to match CLEAN
    master_cols = [
        "file_name", "doi", "title", "authors", "year", "publication_type", "source_category",
        "venue", "open_access", "corpus_source", "json_source_path",
        "student_weights_used", "replicate_weights_used", "weight_variable_name", "weight_fields_interpretation",
        "plausible_values_handling", "missing_data_handling", "handling_not_reported_explanation",
        "research_design_type", "outcome_summary", "null_fields_interpretation",
        "ml_primary", "ml_all_techniques", "total_students", "sample_filtering_criteria",
        "countries_formatted", "countries_json", "effect_size", "primary_finding",
        "document_class", "ml_techniques", "sample_size", "confounders",
        "study_filter_type", "ml_family", "pv_filter_label", "md_filter_label", "weights_filter",
    ]
    for col in master_cols:
        if col not in df_master.columns:
            df_master[col] = ""
    df_master = df_master[[c for c in master_cols if c in df_master.columns]]

    # Write Excel
    print(f"Writing Excel to {OUTPUT_PATH}")
    from openpyxl import Workbook
    wb = Workbook()

    sheet_order = [
        ("0_Dashboard_Analysis_Control", df_dashboard, "tbl_Dashboard"),
        ("1_Articles_Master", df_master, "tbl_ArticlesMaster"),
        ("2_Main_Findings", df_findings, "tbl_MainFindings"),
        ("3_Confounders", df_confounders, "tbl_Confounders"),
    ]

    for i, (name, df, tbl) in enumerate(sheet_order):
        if i == 0:
            ws = wb.active
            ws.title = name
        else:
            ws = wb.create_sheet(name)
        write_df_to_sheet(ws, df)
        format_sheet(ws, tbl)

    wb.save(OUTPUT_PATH)
    print(f"Done. Saved to {OUTPUT_PATH}")
    print(f"  Sheets: {[s[0] for s in sheet_order]}")


if __name__ == "__main__":
    main()
