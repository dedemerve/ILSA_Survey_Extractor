#!/usr/bin/env python3
"""
Build a 3-sheet relational Excel meta-analysis dataset from on-disk article JSON.

Scans ``outputs/**/*.json`` (nested ``json/`` folders included), deduplicates articles
by normalized ``file_name`` then ``doi``, and writes:

  outputs/ILSA_Meta_Analysis_Dataset.xlsx
    - 1_Articles_Master   (one row per article)
    - 2_Main_Findings     (one row per finding)
    - 3_Confounders       (one row per confounder)

  outputs/ILSA_Meta_Analysis_Dataset_CLEAN.xlsx
    Rule-based sanitization + controlled vocabulary (taxonomy filters, dashboard).
    Sheet 0_Dashboard_Analysis_Control — filter counts and Stage 2/3 future-tense note.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import unicodedata
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Any, Iterable

import pandas as pd

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))

from src.enrichment.academic_taxonomy import (
    apply_academic_taxonomy,
    assign_document_class,
    build_analysis_dashboard,
    taxonomy_validation_lists,
    write_classification_audit_log,
)
from src.schemas.models import validate_public_article_json

DEFAULT_OUTPUTS_DIR = PROJECT_ROOT / "outputs"
DEFAULT_XLSX_PATH = DEFAULT_OUTPUTS_DIR / "ILSA_Meta_Analysis_Dataset.xlsx"
DEFAULT_CLEAN_XLSX_PATH = DEFAULT_OUTPUTS_DIR / "ILSA_Meta_Analysis_Dataset_CLEAN.xlsx"
DEFAULT_AUDIT_LOG = DEFAULT_OUTPUTS_DIR / "audit_log.txt"

# ── Sentinel labels (zero imputation: status only, no invented values) ──
SENTINEL_NA_TECHNICAL = "N/A: Technical Report"
SENTINEL_NA_DESCRIPTIVE = "N/A: Descriptive Report"
SENTINEL_NOT_REPORTED_AUTHORS = "Not Reported by Authors"
SENTINEL_NOT_REPORTED_TRADITIONAL = "Not Reported: Likely Traditional Methods"
SENTINEL_NOT_REPORTED = "Not Reported"
SENTINEL_MISSING = "Missing"

_LEGACY_NON_VALUES = frozenset({
    "",
    "not reported",
    "not reported in extraction",
    "unknown",
    "unknown — not in extraction",
    "no country breakdown reported",
    "n/a",
    "na",
    "null",
    "none",
})

# ═══════════════════════════════════════════════════════════════════
# STAGE 2 / STAGE 3 — FUTURE TENSE SAFETY WALL (not implemented yet)
# ═══════════════════════════════════════════════════════════════════
STAGE_FUTURE_TENSE_WARNING = """
PROJECT STAGES NOT YET APPLIED (documentation / generated prose only):
  • Stage 2 — Terminology alignment will be implemented in a future pipeline pass.
  • Stage 3 — RAG-based analytic agent will be proposed and deployed later.

Any manuscript text referring to Stages 2–3 MUST use future tense only
(e.g., "we will implement", "we propose", "planned"), never past tense implying
completion. This script performs Stage-1 rule-based curation only.
"""

_INTERPRETATION_HINTS: dict[str, str] = {
    "ml_techniques": (
        "High N/A among technical reports reflects descriptive IEA/OECD manuals; "
        "empirical gaps often indicate traditional (non-ML) econometric methods."
    ),
    "effect_size": (
        "Technical reports rarely report predictive effect sizes; empirical "
        "articles may omit metrics when authors focus on significance only."
    ),
    "sample_size": (
        "National/framework documents seldom state a single analytic N; "
        "peer-reviewed papers usually report sample size when data are used."
    ),
    "confounders": (
        "Empty confounder lists on technical reports are expected; empirical "
        "articles should list questionnaire controls when regression is run."
    ),
}

# When the same article appears in multiple corpora, keep the higher-priority copy.
_CORPUS_PRIORITY: dict[str, int] = {
    "IEA": 0,
    "OECD": 1,
    "Scopus": 2,
    "Web of Science": 3,
    "ilsa_survey_articles": 4,
}

# ISO 3166-1 alpha-3 → display name (common ILSA participants).
_COUNTRY_NAMES: dict[str, str] = {
    "AUS": "Australia",
    "AUT": "Austria",
    "BEL": "Belgium",
    "BGR": "Bulgaria",
    "BRA": "Brazil",
    "CAN": "Canada",
    "CHE": "Switzerland",
    "CHL": "Chile",
    "CHN": "China",
    "COL": "Colombia",
    "CZE": "Czech Republic",
    "DEU": "Germany",
    "DNK": "Denmark",
    "ESP": "Spain",
    "EST": "Estonia",
    "FIN": "Finland",
    "FRA": "France",
    "GBR": "United Kingdom",
    "GRC": "Greece",
    "HKG": "Hong Kong",
    "HRV": "Croatia",
    "HUN": "Hungary",
    "IDN": "Indonesia",
    "IRL": "Ireland",
    "ISL": "Iceland",
    "ISR": "Israel",
    "ITA": "Italy",
    "JPN": "Japan",
    "KOR": "Korea",
    "LTU": "Lithuania",
    "LUX": "Luxembourg",
    "LVA": "Latvia",
    "MAC": "Macao",
    "MEX": "Mexico",
    "NLD": "Netherlands",
    "NOR": "Norway",
    "NZL": "New Zealand",
    "POL": "Poland",
    "PRT": "Portugal",
    "ROU": "Romania",
    "RUS": "Russia",
    "SGP": "Singapore",
    "SVK": "Slovakia",
    "SVN": "Slovenia",
    "SWE": "Sweden",
    "TUR": "Turkey",
    "TWN": "Chinese Taipei",
    "USA": "United States",
    "ZAF": "South Africa",
}


# openpyxl rejects ASCII control characters in cell strings (common in PDF-derived text).
_ILLEGAL_XLSX_CHARS = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def _sanitize_excel_string(value: Any) -> Any:
    if isinstance(value, str):
        return _ILLEGAL_XLSX_CHARS.sub("", value)
    return value


def _sanitize_rows_for_excel(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [{k: _sanitize_excel_string(v) for k, v in row.items()} for row in rows]


def _is_blank(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, float) and pd.isna(value):
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False


def _fill_blank(series: pd.Series, sentinel: str) -> pd.Series:
    return series.where(~series.apply(_is_blank), sentinel)


def apply_excel_sentinels(
    df_master: pd.DataFrame,
    df_findings: pd.DataFrame,
    df_confounders: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Replace empty Excel cells with explicit sentinels (no fabricated values).
    """
    if not df_master.empty:
        for col in ("doi", "venue", "authors"):
            if col in df_master.columns:
                df_master[col] = _fill_blank(
                    df_master[col], "UNKNOWN — not in extraction"
                )
        if "open_access" in df_master.columns:
            df_master["open_access"] = df_master["open_access"].apply(
                lambda v: "UNKNOWN"
                if _is_blank(v)
                else ("TRUE" if v is True else "FALSE" if v is False else str(v))
            )
        for col in ("student_weights_used", "replicate_weights_used"):
            if col in df_master.columns:
                df_master[col] = df_master[col].apply(
                    lambda v: "UNKNOWN"
                    if _is_blank(v)
                    else ("TRUE" if v is True else "FALSE" if v is False else str(v))
                )
        if "countries_formatted" in df_master.columns:
            df_master["countries_formatted"] = _fill_blank(
                df_master["countries_formatted"],
                "No country breakdown reported",
            )
        if "ml_primary" in df_master.columns and "ml_all_techniques" in df_master.columns:
            def _ml_primary(row: pd.Series) -> str:
                if not _is_blank(row.get("ml_primary")):
                    return str(row["ml_primary"])
                techniques = row.get("ml_all_techniques")
                if isinstance(techniques, str) and techniques.strip():
                    first = techniques.split(",")[0].strip()
                    if first:
                        return first
                return "Not reported in extraction"

            df_master["ml_primary"] = df_master.apply(_ml_primary, axis=1)
        for col in (
            "research_design_type",
            "publication_type",
            "source_category",
            "weight_variable_name",
        ):
            if col in df_master.columns:
                df_master[col] = _fill_blank(df_master[col], "Not reported in extraction")
        if "total_students" in df_master.columns:
            df_master["total_students"] = _fill_blank(
                df_master["total_students"], "Not reported in extraction"
            )

    if not df_findings.empty and "performance_metrics" in df_findings.columns:
        df_findings["performance_metrics"] = _fill_blank(
            df_findings["performance_metrics"], "Not reported"
        )
        for col in ("dataset_used", "target_variable", "standardized_conclusion", "top_predictors"):
            if col in df_findings.columns:
                df_findings[col] = _fill_blank(df_findings[col], "Not reported in extraction")

    if not df_confounders.empty:
        for col in ("variable_code", "variable_name", "category"):
            if col in df_confounders.columns:
                df_confounders[col] = _fill_blank(
                    df_confounders[col], "Not reported in extraction"
                )

    return df_master, df_findings, df_confounders


def _is_legacy_blank(value: Any) -> bool:
    if _is_blank(value):
        return True
    if isinstance(value, str):
        return value.strip().lower() in _LEGACY_NON_VALUES
    return False


def _assign_document_class(row: pd.Series) -> str:
    """Delegates to academic_taxonomy (word-boundary safe; Soru A/D)."""
    return assign_document_class(row)


def _apply_conditional_fill(
    series: pd.Series,
    doc_class: pd.Series,
    *,
    technical_value: str,
    empirical_value: str,
    unclassified_value: str = SENTINEL_MISSING,
) -> pd.Series:
    out = series.astype(object).copy()
    dc = doc_class.reset_index(drop=True)
    out = out.reset_index(drop=True)

    for i in range(len(out)):
        if not _is_legacy_blank(out.iat[i]):
            continue
        label = dc.iat[i] if i < len(dc) else "unclassified"
        if label == "technical_report":
            out.iat[i] = technical_value
        elif label == "empirical_article":
            out.iat[i] = empirical_value
        else:
            out.iat[i] = unclassified_value
    return out


def _merge_first_finding_metrics(
    df_master: pd.DataFrame,
    df_findings: pd.DataFrame,
) -> pd.DataFrame:
    """Add effect_size and primary_finding on master from exploded findings."""
    master = df_master.copy()
    for drop_col in ("effect_size", "primary_finding"):
        if drop_col in master.columns:
            master = master.drop(columns=[drop_col])

    if df_findings.empty or "file_name" not in df_findings.columns:
        master["effect_size"] = None
        master["primary_finding"] = (
            master["outcome_summary"] if "outcome_summary" in master.columns else None
        )
        return master

    agg = df_findings.groupby("file_name", dropna=False).first().reset_index()
    if "performance_metrics" in agg.columns:
        master = master.merge(
            agg[["file_name", "performance_metrics"]].rename(
                columns={"performance_metrics": "effect_size"}
            ),
            on="file_name",
            how="left",
        )
    else:
        master["effect_size"] = None

    if "standardized_conclusion" in agg.columns:
        master = master.merge(
            agg[["file_name", "standardized_conclusion"]].rename(
                columns={"standardized_conclusion": "primary_finding"}
            ),
            on="file_name",
            how="left",
        )
    else:
        master["primary_finding"] = None

    if "outcome_summary" in master.columns:
        mask = master["primary_finding"].apply(_is_legacy_blank)
        master.loc[mask, "primary_finding"] = master.loc[mask, "outcome_summary"]

    return master


def _attach_document_class(
    df: pd.DataFrame,
    master_lookup: pd.DataFrame,
) -> pd.DataFrame:
    if df.empty or "file_name" not in df.columns:
        return df
    out = df.copy()
    for col in ("publication_type", "source_category", "document_class"):
        if col in out.columns:
            out = out.drop(columns=[col])
    keys = master_lookup[
        ["file_name", "publication_type", "source_category", "document_class"]
    ].drop_duplicates(subset=["file_name"])
    return out.merge(keys, on="file_name", how="left")


def _ensure_relational_placeholders(
    df_master: pd.DataFrame,
    df_findings: pd.DataFrame,
    df_confounders: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """One placeholder row per article missing from findings/confounders (explosion context)."""
    findings = df_findings.copy() if not df_findings.empty else pd.DataFrame()
    confounders = df_confounders.copy() if not df_confounders.empty else pd.DataFrame()

    if df_master.empty:
        return findings, confounders

    master_keys = df_master[
        ["file_name", "doi", "publication_type", "source_category", "document_class"]
    ].copy()

    finding_files = (
        set(findings["file_name"].dropna().unique()) if "file_name" in findings.columns else set()
    )
    conf_files = (
        set(confounders["file_name"].dropna().unique())
        if "file_name" in confounders.columns
        else set()
    )

    new_findings: list[dict[str, Any]] = []
    new_confounders: list[dict[str, Any]] = []

    for _, row in master_keys.iterrows():
        fn = row["file_name"]
        if fn not in finding_files:
            new_findings.append(
                {
                    "file_name": fn,
                    "doi": row.get("doi"),
                    "dataset_used": None,
                    "target_variable": None,
                    "top_predictors": None,
                    "performance_metrics": None,
                    "standardized_conclusion": None,
                    "publication_type": row.get("publication_type"),
                    "source_category": row.get("source_category"),
                    "document_class": row.get("document_class"),
                }
            )
        if fn not in conf_files:
            new_confounders.append(
                {
                    "file_name": fn,
                    "doi": row.get("doi"),
                    "variable_code": None,
                    "variable_name": None,
                    "category": None,
                    "description": None,
                    "publication_type": row.get("publication_type"),
                    "source_category": row.get("source_category"),
                    "document_class": row.get("document_class"),
                }
            )

    if new_findings:
        findings = pd.concat([findings, pd.DataFrame(new_findings)], ignore_index=True)
    if new_confounders:
        confounders = pd.concat([confounders, pd.DataFrame(new_confounders)], ignore_index=True)

    return findings, confounders


def categorize_missing_values(
    df_master: pd.DataFrame,
    df_findings: pd.DataFrame,
    df_confounders: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Rule-based conditional sanitization for heterogeneous ILSA literature.

    Labels empty cells as N/A (technical), Not Reported (empirical), or Missing
    (unclassified) — never invents substantive values (zero imputation).
    """
    master = df_master.copy()
    findings = df_findings.copy()
    confounders = df_confounders.copy()

    if master.empty:
        return master, findings, confounders

    master = _merge_first_finding_metrics(master, findings)
    master["document_class"] = master.apply(_assign_document_class, axis=1)
    doc_class = master["document_class"]

    if "ml_all_techniques" in master.columns:
        master["ml_techniques"] = master.apply(
            lambda r: r["ml_primary"]
            if not _is_legacy_blank(r.get("ml_primary"))
            else r.get("ml_all_techniques"),
            axis=1,
        )
    elif "ml_primary" in master.columns:
        master["ml_techniques"] = master["ml_primary"]

    master["sample_size"] = master["total_students"] if "total_students" in master.columns else None

    master["ml_techniques"] = _apply_conditional_fill(
        master["ml_techniques"] if "ml_techniques" in master.columns else pd.Series([None] * len(master)),
        doc_class,
        technical_value=SENTINEL_NA_TECHNICAL,
        empirical_value=SENTINEL_NOT_REPORTED_TRADITIONAL,
    )
    master["effect_size"] = _apply_conditional_fill(
        master["effect_size"] if "effect_size" in master.columns else pd.Series([None] * len(master)),
        doc_class,
        technical_value=SENTINEL_NA_TECHNICAL,
        empirical_value=SENTINEL_NOT_REPORTED_AUTHORS,
    )
    master["sample_size"] = _apply_conditional_fill(
        master["sample_size"] if "sample_size" in master.columns else pd.Series([None] * len(master)),
        doc_class,
        technical_value=SENTINEL_NA_TECHNICAL,
        empirical_value=SENTINEL_NOT_REPORTED,
    )
    master["primary_finding"] = _apply_conditional_fill(
        master["primary_finding"] if "primary_finding" in master.columns else pd.Series([None] * len(master)),
        doc_class,
        technical_value=SENTINEL_NA_DESCRIPTIVE,
        empirical_value=SENTINEL_NOT_REPORTED_AUTHORS,
    )

    conf_col = (
        "variable_name"
        if "variable_name" in confounders.columns
        else "variable_code"
    )
    if not confounders.empty and conf_col in confounders.columns:
        conf_has_value = confounders.groupby("file_name")[conf_col].apply(
            lambda s: s.apply(lambda v: not _is_legacy_blank(v)).any()
        )
        master["confounders"] = master["file_name"].map(conf_has_value).map(
            {True: "present", False: None}
        )
    else:
        master["confounders"] = None

    master["confounders"] = _apply_conditional_fill(
        master["confounders"],
        doc_class,
        technical_value=SENTINEL_NA_TECHNICAL,
        empirical_value=SENTINEL_NOT_REPORTED_AUTHORS,
    )

    master_lookup = master
    findings, confounders = _ensure_relational_placeholders(master, findings, confounders)
    findings = _attach_document_class(findings, master_lookup)
    confounders = _attach_document_class(confounders, master_lookup)

    if "document_class" not in findings.columns:
        findings["document_class"] = "unclassified"
    f_doc = findings["document_class"]

    if "performance_metrics" in findings.columns:
        findings["effect_size"] = findings["performance_metrics"]
        findings["effect_size"] = _apply_conditional_fill(
            findings["effect_size"],
            f_doc,
            technical_value=SENTINEL_NA_TECHNICAL,
            empirical_value=SENTINEL_NOT_REPORTED_AUTHORS,
        )
    if "standardized_conclusion" in findings.columns:
        findings["primary_finding"] = findings["standardized_conclusion"]
        findings["primary_finding"] = _apply_conditional_fill(
            findings["primary_finding"],
            f_doc,
            technical_value=SENTINEL_NA_DESCRIPTIVE,
            empirical_value=SENTINEL_NOT_REPORTED_AUTHORS,
        )
    for col in ("dataset_used", "target_variable", "top_predictors"):
        if col in findings.columns:
            findings[col] = _apply_conditional_fill(
                findings[col],
                f_doc,
                technical_value=SENTINEL_NA_TECHNICAL,
                empirical_value=SENTINEL_NOT_REPORTED_AUTHORS,
            )

    if not confounders.empty:
        if "document_class" not in confounders.columns:
            confounders["document_class"] = "unclassified"
        c_doc = confounders["document_class"]
        for col in ("variable_code", "variable_name", "category", "description"):
            if col in confounders.columns:
                confounders[col] = _apply_conditional_fill(
                    confounders[col],
                    c_doc,
                    technical_value=SENTINEL_NA_TECHNICAL,
                    empirical_value=SENTINEL_NOT_REPORTED_AUTHORS,
                )

    return master, findings, confounders


def _is_structured_missing(value: Any) -> bool:
    """True when cell has no substantive extracted value (blank or sentinel label)."""
    if _is_legacy_blank(value):
        return True
    if isinstance(value, str):
        low = value.strip().lower()
        if low == "present":
            return False
        if low.startswith("n/a:") or low.startswith("not reported") or low == "missing":
            return True
    return False


def _missing_rate(series: pd.Series) -> float:
    if series.empty:
        return 0.0
    return 100.0 * series.apply(_is_structured_missing).sum() / len(series)


def generate_missing_value_table(
    df_master: pd.DataFrame,
    df_findings: pd.DataFrame,
    df_confounders: pd.DataFrame,
) -> str:
    """Build LaTeX table of missing-value rates by document class (for manuscript)."""
    if "document_class" not in df_master.columns:
        df_master = df_master.copy()
        df_master["document_class"] = df_master.apply(_assign_document_class, axis=1)

    tech = df_master[df_master["document_class"] == "technical_report"]
    emp = df_master[df_master["document_class"] == "empirical_article"]

    if not df_findings.empty and "document_class" in df_findings.columns:
        tech_f = df_findings[df_findings["document_class"] == "technical_report"]
        emp_f = df_findings[df_findings["document_class"] == "empirical_article"]
    else:
        tech_f = pd.DataFrame()
        emp_f = pd.DataFrame()

    if not df_confounders.empty and "document_class" in df_confounders.columns:
        tech_c = df_confounders[df_confounders["document_class"] == "technical_report"]
        emp_c = df_confounders[df_confounders["document_class"] == "empirical_article"]
    else:
        tech_c = pd.DataFrame()
        emp_c = pd.DataFrame()

    def _rate_master(col: str, subset: pd.DataFrame) -> float:
        if col not in subset.columns:
            return 0.0
        return _missing_rate(subset[col])

    def _rate_findings(col: str, subset: pd.DataFrame) -> float:
        if subset.empty or col not in subset.columns:
            return 0.0
        return _missing_rate(subset[col])

    def _rate_confounders(subset: pd.DataFrame) -> float:
        if subset.empty:
            return 100.0
        col = "variable_name" if "variable_name" in subset.columns else "variable_code"
        if col not in subset.columns:
            return 0.0
        return _missing_rate(subset[col])

    rows_spec = [
        (
            "ml\\_techniques",
            "ml_techniques",
            _rate_master("ml_techniques", tech),
            _rate_master("ml_techniques", emp),
            "ml_techniques",
        ),
        (
            "effect\\_size",
            "effect_size",
            _rate_findings("effect_size", tech_f) if not tech_f.empty else _rate_master("effect_size", tech),
            _rate_findings("effect_size", emp_f) if not emp_f.empty else _rate_master("effect_size", emp),
            "effect_size",
        ),
        (
            "sample\\_size",
            "sample_size",
            _rate_master("sample_size", tech),
            _rate_master("sample_size", emp),
            "sample_size",
        ),
        (
            "confounders",
            "confounders",
            _rate_confounders(tech_c) if not tech_c.empty else _rate_master("confounders", tech),
            _rate_confounders(emp_c) if not emp_c.empty else _rate_master("confounders", emp),
            "confounders",
        ),
    ]

    body_lines: list[str] = []
    for label, _key, pct_tech, pct_emp, hint_key in rows_spec:
        interp = _INTERPRETATION_HINTS.get(hint_key, "See tablenotes.")
        body_lines.append(
            f"{label} & {pct_tech:.1f}\\% & {pct_emp:.1f}\\% & {interp} \\\\"
        )

    n_tech = len(tech)
    n_emp = len(emp)
    notes = (
        f"\\\\[0.5em]\\footnotesize\\textit{{Notes.}} "
        f"Rates are shares of non-substantive cells (blank, legacy placeholder, or sentinel), "
        f"among $N={n_tech}$ technical reports and $N={n_emp}$ empirical articles "
        f"(unique master rows). Stage~2 terminology alignment and Stage~3 RAG analytics "
        f"\\textbf{{will be implemented}} in future work; this table reflects Stage~1 "
        f"curation only. No values were imputed."
    )

    latex = (
        "\\begin{table}[htbp]\n"
        "\\caption{Missing Value Profile by Document Type}\n"
        "\\label{tab:missing_profile}\n"
        "\\centering\n"
        "\\small\n"
        "\\begin{tabular}{@{}lrrp{6.2cm}@{}}\n"
        "\\toprule\n"
        "\\textbf{Field} & \\textbf{Technical Reports} & "
        "\\textbf{Empirical Articles} & \\textbf{Interpretation} \\\\\n"
        "\\midrule\n"
        + "\n".join(body_lines)
        + "\n\\bottomrule\n"
        f"\\end{{tabular}}\n"
        f"{notes}\n"
        "\\end{table}\n"
    )
    return latex


def read_excel_dataset(xlsx_path: Path) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Load the three relational sheets from an existing workbook."""
    master = pd.read_excel(xlsx_path, sheet_name="1_Articles_Master")
    findings = pd.read_excel(xlsx_path, sheet_name="2_Main_Findings")
    confounders = pd.read_excel(xlsx_path, sheet_name="3_Confounders")
    return master, findings, confounders


def _apply_excel_filter_dropdowns(xlsx_path: Path) -> None:
    """Attach list validations for taxonomy columns (Excel filter dropdowns)."""
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    lists = taxonomy_validation_lists()
    sheet_columns: dict[str, list[str]] = {
        "1_Articles_Master": [
            "study_filter_type",
            "ml_family",
            "pv_filter_label",
            "md_filter_label",
            "weights_filter",
        ],
        "2_Main_Findings": [
            "study_filter_type",
            "target_domain",
            "target_dimension",
        ],
        "3_Confounders": [
            "study_filter_type",
            "predictor_level",
            "predictor_category",
        ],
    }

    wb = load_workbook(xlsx_path)
    for sheet_name, cols in sheet_columns.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        if ws.max_row < 2:
            continue
        header = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
        for col_name in cols:
            if col_name not in lists or col_name not in header:
                continue
            col_idx = header[col_name]
            col_letter = get_column_letter(col_idx)
            quoted = ",".join(lists[col_name])
            dv = DataValidation(
                type="list",
                formula1=f'"{quoted}"',
                allow_blank=True,
                showErrorMessage=True,
                errorTitle="Invalid filter value",
                error="Choose a value from the controlled vocabulary list.",
            )
            dv.add(f"{col_letter}2:{col_letter}{ws.max_row}")
            ws.add_data_validation(dv)
    wb.save(xlsx_path)


def write_clean_excel(
    df_master: pd.DataFrame,
    df_findings: pd.DataFrame,
    df_confounders: pd.DataFrame,
    output_path: Path,
    *,
    df_dashboard: pd.DataFrame | None = None,
    add_dropdowns: bool = True,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    df_master = pd.DataFrame(_sanitize_rows_for_excel(df_master.to_dict(orient="records")))
    df_findings = pd.DataFrame(_sanitize_rows_for_excel(df_findings.to_dict(orient="records")))
    df_confounders = pd.DataFrame(
        _sanitize_rows_for_excel(df_confounders.to_dict(orient="records"))
    )
    if df_dashboard is None:
        df_dashboard = build_analysis_dashboard(df_master, df_findings, df_confounders)
    df_dashboard = pd.DataFrame(_sanitize_rows_for_excel(df_dashboard.to_dict(orient="records")))

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df_dashboard.to_excel(writer, sheet_name="0_Dashboard_Analysis_Control", index=False)
        df_master.to_excel(writer, sheet_name="1_Articles_Master", index=False)
        df_findings.to_excel(writer, sheet_name="2_Main_Findings", index=False)
        df_confounders.to_excel(writer, sheet_name="3_Confounders", index=False)

    if add_dropdowns:
        _apply_excel_filter_dropdowns(output_path)


def build_clean_dataset(
    df_master: pd.DataFrame,
    df_findings: pd.DataFrame,
    df_confounders: pd.DataFrame,
    output_path: Path,
    *,
    print_latex: bool = True,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Apply conditional sanitization, optional LaTeX table, save CLEAN workbook."""
    print(STAGE_FUTURE_TENSE_WARNING.strip())
    clean_master, clean_findings, clean_confounders = categorize_missing_values(
        df_master, df_findings, df_confounders
    )
    clean_master, clean_findings, clean_confounders = apply_academic_taxonomy(
        clean_master, clean_findings, clean_confounders
    )
    n_conflicts = write_classification_audit_log(clean_master, DEFAULT_AUDIT_LOG.resolve())
    if n_conflicts:
        print(
            f"WARNING: {n_conflicts} journal/technical_report conflict(s) logged to "
            f"outputs/audit_log.txt",
            flush=True,
        )
    else:
        print("Classification audit: 0 journal/technical_report conflicts.", flush=True)
    dashboard = build_analysis_dashboard(
        clean_master, clean_findings, clean_confounders
    )
    if print_latex:
        print("\n% --- LaTeX: Missing Value Profile (paste into manuscript) ---\n")
        print(generate_missing_value_table(clean_master, clean_findings, clean_confounders))
    write_clean_excel(
        clean_master,
        clean_findings,
        clean_confounders,
        output_path,
        df_dashboard=dashboard,
    )
    print(f"\nWrote {output_path}")
    return clean_master, clean_findings, clean_confounders


def _norm_file_name(value: str | None) -> str:
    if not value:
        return ""
    text = unicodedata.normalize("NFKD", value)
    text = text.encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]+", "", text.lower())


def _norm_doi(value: str | None) -> str:
    if not value:
        return ""
    doi = str(value).strip().lower()
    for prefix in (
        "https://doi.org/",
        "http://doi.org/",
        "https://dx.doi.org/",
        "http://dx.doi.org/",
    ):
        if doi.startswith(prefix):
            doi = doi[len(prefix) :]
    return doi


def _join_list(values: Any, sep: str = ", ") -> str | None:
    if values is None:
        return None
    if isinstance(values, str):
        return values.strip() or None
    if isinstance(values, list):
        parts = [str(v).strip() for v in values if v is not None and str(v).strip()]
        return sep.join(parts) if parts else None
    return str(values)


def _country_label(code: str) -> str:
    code = (code or "").strip().upper()
    return _COUNTRY_NAMES.get(code, code)


def _format_countries(countries: Any) -> str | None:
    if not countries:
        return None
    parts: list[str] = []
    for item in countries:
        if not isinstance(item, dict):
            continue
        code = str(item.get("country_code") or "").strip()
        n = item.get("n_students")
        label = _country_label(code) if code else "Unknown"
        if n is not None and str(n).strip() != "":
            parts.append(f"{label} ({n})")
        else:
            parts.append(label)
    return ", ".join(parts) if parts else None


def _corpus_from_path(json_path: Path, outputs_dir: Path) -> str:
    try:
        return json_path.relative_to(outputs_dir).parts[0]
    except ValueError:
        return "unknown"


def _corpus_rank(corpus: str) -> int:
    return _CORPUS_PRIORITY.get(corpus, 99)


def _is_article_json(raw: object) -> bool:
    return isinstance(raw, dict) and "metadata" in raw and "data" in raw


def _dedupe_key(meta: dict[str, Any]) -> tuple[str, str]:
    return (_norm_file_name(meta.get("file_name")), _norm_doi(meta.get("doi")))


def _pick_preferred(
    current: tuple[Path, dict[str, Any], str],
    candidate: tuple[Path, dict[str, Any], str],
) -> tuple[Path, dict[str, Any], str]:
    cur_path, cur_raw, cur_corpus = current
    cand_path, cand_raw, cand_corpus = candidate
    if _corpus_rank(cand_corpus) < _corpus_rank(cur_corpus):
        return candidate
    if _corpus_rank(cand_corpus) > _corpus_rank(cur_corpus):
        return current
    # Same corpus tier: prefer longer JSON path (more specific nested folder).
    if len(str(cand_path)) > len(str(cur_path)):
        return candidate
    return current


def discover_json_paths(outputs_dir: Path) -> list[Path]:
    return sorted(outputs_dir.rglob("*.json"))


def _load_json_file(path: Path) -> tuple[Path, dict[str, Any] | None, str | None]:
    """Load one JSON file; returns (path, raw_dict_or_none, error_message_or_none)."""
    try:
        raw = json.loads(path.read_text(encoding="utf-8"))
        if not isinstance(raw, dict):
            return path, None, "not a JSON object"
        return path, raw, None
    except (OSError, json.JSONDecodeError) as exc:
        return path, None, str(exc)


def load_and_dedupe_articles(
    json_paths: Iterable[Path],
    outputs_dir: Path,
    *,
    dedupe: bool = True,
    validate: bool = False,
    verbose: bool = False,
    workers: int = 8,
) -> tuple[list[dict[str, Any]], list[str]]:
    """Return deduplicated article dicts and warning lines."""
    by_key: dict[tuple[str, str], tuple[Path, dict[str, Any], str]] = {}
    fn_index: dict[str, tuple[str, str]] = {}
    doi_index: dict[str, tuple[str, str]] = {}
    all_entries: list[tuple[Path, dict[str, Any], str]] = []
    warnings: list[str] = []
    skipped = 0
    path_list = list(json_paths)
    max_workers = max(1, min(workers, len(path_list) or 1))

    loaded: list[tuple[Path, dict[str, Any]]] = []
    if max_workers == 1:
        for idx, path in enumerate(path_list, start=1):
            if verbose and (idx == 1 or idx % 200 == 0 or idx == len(path_list)):
                print(f"  Reading JSON {idx}/{len(path_list)} …", flush=True)
            path, raw, err = _load_json_file(path)
            if err is not None:
                warnings.append(f"SKIP (read): {path} — {err}")
                skipped += 1
                continue
            assert raw is not None
            loaded.append((path, raw))
    else:
        done = 0
        with ThreadPoolExecutor(max_workers=max_workers) as ex:
            futures = {ex.submit(_load_json_file, p): p for p in path_list}
            for fut in as_completed(futures):
                done += 1
                if verbose and (done == 1 or done % 200 == 0 or done == len(path_list)):
                    print(f"  Reading JSON {done}/{len(path_list)} …", flush=True)
                path, raw, err = fut.result()
                if err is not None:
                    warnings.append(f"SKIP (read): {path} — {err}")
                    skipped += 1
                    continue
                assert raw is not None
                loaded.append((path, raw))

    for path, raw in loaded:
        if not _is_article_json(raw):
            warnings.append(f"SKIP (not article schema): {path}")
            skipped += 1
            continue

        if validate:
            try:
                model = validate_public_article_json(raw)
                raw = model.model_dump(mode="python")
            except Exception as exc:
                warnings.append(f"SKIP (validation): {path} — {exc}")
                skipped += 1
                continue

        corpus = _corpus_from_path(path, outputs_dir)
        meta = raw.get("metadata") or {}
        key = _dedupe_key(meta)

        if not dedupe:
            all_entries.append((path, raw, corpus))
            continue

        norm_fn, norm_doi = key
        if not norm_fn and not norm_doi:
            unique_key = (f"__path__:{path}", "")
            by_key[unique_key] = (path, raw, corpus)
            continue

        existing_key: tuple[str, str] | None = None
        if norm_fn and norm_fn in fn_index:
            existing_key = fn_index[norm_fn]
        elif norm_doi and norm_doi in doi_index:
            existing_key = doi_index[norm_doi]

        if existing_key is None:
            new_key = (norm_fn, norm_doi)
            by_key[new_key] = (path, raw, corpus)
            if norm_fn:
                fn_index[norm_fn] = new_key
            if norm_doi:
                doi_index[norm_doi] = new_key
        else:
            by_key[existing_key] = _pick_preferred(by_key[existing_key], (path, raw, corpus))
            if norm_fn:
                fn_index[norm_fn] = existing_key
            if norm_doi:
                doi_index[norm_doi] = existing_key

    articles: list[dict[str, Any]] = []
    entries = all_entries if not dedupe else list(by_key.values())

    for path, raw, corpus in entries:
        meta = raw.get("metadata") or {}
        data = raw.get("data") or {}
        articles.append(
            {
                "_json_path": str(path),
                "_corpus": corpus,
                "metadata": meta,
                "data": data,
            }
        )

    warnings.append(
        f"Loaded {len(path_list)} JSON file(s); "
        f"{len(articles)} unique article(s); skipped {skipped}."
    )
    return articles, warnings


def build_master_rows(articles: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for article in articles:
        meta = article.get("metadata") or {}
        data = article.get("data") or {}
        survey = data.get("survey_design") or {}
        sample = data.get("sample_details") or {}
        ml = data.get("ml_techniques") or {}

        file_name = meta.get("file_name")
        rows.append(
            {
                "file_name": file_name,
                "doi": meta.get("doi"),
                "title": meta.get("title"),
                "authors": _join_list(meta.get("authors")),
                "year": meta.get("year"),
                "publication_type": meta.get("publication_type"),
                "source_category": meta.get("source_category"),
                "venue": meta.get("venue"),
                "open_access": meta.get("open_access"),
                "corpus_source": article.get("_corpus"),
                "json_source_path": article.get("_json_path"),
                "student_weights_used": survey.get("student_weights_used"),
                "replicate_weights_used": survey.get("replicate_weights_used"),
                "weight_variable_name": survey.get("weight_variable_name"),
                "weight_fields_interpretation": survey.get("weight_fields_interpretation"),
                "plausible_values_handling": data.get("plausible_values_handling"),
                "missing_data_handling": data.get("missing_data_handling"),
                "handling_not_reported_explanation": data.get(
                    "handling_not_reported_explanation"
                ),
                "research_design_type": data.get("research_design_type"),
                "outcome_summary": data.get("outcome_summary"),
                "null_fields_interpretation": data.get("null_fields_interpretation"),
                "ml_primary": ml.get("primary"),
                "ml_all_techniques": _join_list(ml.get("all_techniques")),
                "total_students": sample.get("total_students"),
                "sample_filtering_criteria": sample.get("sample_filtering_criteria"),
                "countries_formatted": _format_countries(sample.get("countries")),
                "countries_json": json.dumps(
                    sample.get("countries"), ensure_ascii=False
                )
                if sample.get("countries")
                else None,
            }
        )
    return rows


def build_findings_rows(articles: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for article in articles:
        meta = article.get("metadata") or {}
        data = article.get("data") or {}
        file_name = meta.get("file_name")
        findings = data.get("main_findings") or []
        if not isinstance(findings, list):
            continue
        for finding in findings:
            if not isinstance(finding, dict):
                continue
            rows.append(
                {
                    "file_name": file_name,
                    "doi": meta.get("doi"),
                    "dataset_used": finding.get("dataset_used"),
                    "target_variable": finding.get("target_variable"),
                    "top_predictors": _join_list(finding.get("top_predictors")),
                    "performance_metrics": finding.get("performance_metrics"),
                    "standardized_conclusion": finding.get("standardized_conclusion"),
                }
            )
    return rows


def build_confounders_rows(articles: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for article in articles:
        meta = article.get("metadata") or {}
        data = article.get("data") or {}
        file_name = meta.get("file_name")
        confounders = data.get("confounders_identified") or []
        if not isinstance(confounders, list):
            continue
        for conf in confounders:
            if not isinstance(conf, dict):
                continue
            rows.append(
                {
                    "file_name": file_name,
                    "doi": meta.get("doi"),
                    "variable_code": conf.get("variable_code"),
                    "variable_name": conf.get("variable_name"),
                    "category": conf.get("category"),
                    # Schema has no description field; reserved for forward compatibility.
                    "description": conf.get("description"),
                }
            )
    return rows


def write_excel(
    master_rows: list[dict[str, Any]],
    findings_rows: list[dict[str, Any]],
    confounders_rows: list[dict[str, Any]],
    output_path: Path,
    *,
    apply_legacy_sentinels: bool = True,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    df_master = pd.DataFrame(_sanitize_rows_for_excel(master_rows))
    df_findings = pd.DataFrame(_sanitize_rows_for_excel(findings_rows))
    df_confounders = pd.DataFrame(_sanitize_rows_for_excel(confounders_rows))
    if apply_legacy_sentinels:
        df_master, df_findings, df_confounders = apply_excel_sentinels(
            df_master, df_findings, df_confounders
        )

    with pd.ExcelWriter(
        output_path,
        engine="openpyxl",
    ) as writer:
        df_master.to_excel(writer, sheet_name="1_Articles_Master", index=False)
        df_findings.to_excel(writer, sheet_name="2_Main_Findings", index=False)
        df_confounders.to_excel(writer, sheet_name="3_Confounders", index=False)

    return df_master, df_findings, df_confounders


def build_dataset(
    outputs_dir: Path,
    output_path: Path,
    *,
    clean_output_path: Path | None = DEFAULT_CLEAN_XLSX_PATH,
    dedupe: bool = True,
    validate: bool = False,
    verbose: bool = False,
    workers: int = 8,
) -> dict[str, int]:
    paths = discover_json_paths(outputs_dir)
    print(f"Discovered {len(paths)} JSON file(s) under {outputs_dir}", flush=True)
    articles, warnings = load_and_dedupe_articles(
        paths,
        outputs_dir,
        dedupe=dedupe,
        validate=validate,
        verbose=verbose,
        workers=workers,
    )

    master_rows = build_master_rows(articles)
    findings_rows = build_findings_rows(articles)
    confounders_rows = build_confounders_rows(articles)

    df_master, df_findings, df_confounders = write_excel(
        master_rows,
        findings_rows,
        confounders_rows,
        output_path,
        apply_legacy_sentinels=True,
    )

    if clean_output_path is not None:
        df_raw_master = pd.DataFrame(_sanitize_rows_for_excel(master_rows))
        df_raw_findings = pd.DataFrame(_sanitize_rows_for_excel(findings_rows))
        df_raw_confounders = pd.DataFrame(_sanitize_rows_for_excel(confounders_rows))
        build_clean_dataset(
            df_raw_master,
            df_raw_findings,
            df_raw_confounders,
            clean_output_path.resolve(),
        )

    for line in warnings:
        print(line)

    stats = {
        "json_files_scanned": len(paths),
        "articles_master": len(master_rows),
        "main_findings_rows": len(findings_rows),
        "confounders_rows": len(confounders_rows),
    }
    print(
        f"Wrote {output_path}\n"
        f"  1_Articles_Master: {stats['articles_master']} rows\n"
        f"  2_Main_Findings:   {stats['main_findings_rows']} rows\n"
        f"  3_Confounders:     {stats['confounders_rows']} rows"
    )
    return stats


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--outputs-dir",
        type=Path,
        default=DEFAULT_OUTPUTS_DIR,
        help=f"Root directory to rglob for article JSON (default: {DEFAULT_OUTPUTS_DIR})",
    )
    parser.add_argument(
        "--output-xlsx",
        type=Path,
        default=DEFAULT_XLSX_PATH,
        help=f"Destination Excel path (default: {DEFAULT_XLSX_PATH})",
    )
    parser.add_argument(
        "--clean-output-xlsx",
        type=Path,
        default=DEFAULT_CLEAN_XLSX_PATH,
        help=f"Conditional-sanitization workbook (default: {DEFAULT_CLEAN_XLSX_PATH})",
    )
    parser.add_argument(
        "--clean-only",
        type=Path,
        metavar="XLSX",
        nargs="?",
        const=DEFAULT_XLSX_PATH,
        help="Skip JSON build; read existing workbook and write CLEAN + LaTeX",
    )
    parser.add_argument(
        "--no-clean",
        action="store_true",
        help="Do not write the CLEAN workbook or print the LaTeX table.",
    )
    parser.add_argument(
        "--no-latex",
        action="store_true",
        help="Skip printing the LaTeX missing-value table.",
    )
    parser.add_argument(
        "--no-dedupe",
        action="store_true",
        help="Keep every JSON file as a separate article row (not recommended).",
    )
    parser.add_argument(
        "--validate",
        action="store_true",
        help="Validate each JSON with Pydantic before export (slower).",
    )
    parser.add_argument(
        "-v",
        "--verbose",
        action="store_true",
        help="Print progress while reading JSON files.",
    )
    parser.add_argument(
        "--workers",
        type=int,
        default=8,
        help="Parallel JSON readers (default: 8). Use 1 for deterministic single-threaded I/O.",
    )
    args = parser.parse_args()

    if args.clean_only is not None:
        source = Path(args.clean_only).resolve()
        master, findings, confounders = read_excel_dataset(source)
        build_clean_dataset(
            master,
            findings,
            confounders,
            args.clean_output_xlsx.resolve(),
            print_latex=not args.no_latex,
        )
        return

    build_dataset(
        args.outputs_dir.resolve(),
        args.output_xlsx.resolve(),
        clean_output_path=None if args.no_clean else args.clean_output_xlsx.resolve(),
        dedupe=not args.no_dedupe,
        validate=args.validate,
        verbose=args.verbose,
        workers=args.workers,
    )


if __name__ == "__main__":
    main()
